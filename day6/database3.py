'''from openpyxl import Workbook
wb = Workbook()
sheet = wb.active
data = [
    ('id','wname','year','status','company'),
    (1002,"Java",2020,1,"ITC"),
    (1003,"IoT",2018,1,"Amazon")
]
for row in data:
    sheet.append(row)

wb.save("student.xlsx")
'''
'''for loading'''
from openpyxl import Workbook
import openpyxl as pyxl

wb = pyxl.load_workbook("student.xlsx")
sheet = wb.active
for row in sheet.iter_rows():
    data = [c. value for c in row]
    print(data)